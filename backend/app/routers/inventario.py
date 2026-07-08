from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import load_workbook
from pymongo import ReturnDocument

from app.core.auditoria import registar
from app.core.database import get_database
from app.core.deps import get_current_catequista
from app.core.mongo_utils import object_id_or_404
from app.models.auditoria import AcaoAuditoria
from app.models.catequisando import ErroImportacao, ImportacaoResultado
from app.models.catequista import CatequistaOut
from app.models.inventario import EstadoItem, ItemInventarioCreate, ItemInventarioOut, ItemInventarioUpdate

router = APIRouter(prefix="/inventario", tags=["inventário"])

_MAPA_ESTADO = {
    "bom": EstadoItem.BOM,
    "bom estado": EstadoItem.BOM,
    "em uso": EstadoItem.EM_USO,
    "novo": EstadoItem.NOVO,
    "descartado": EstadoItem.DESCARTADO,
    "danificado": EstadoItem.DANIFICADO,
    "antigo": EstadoItem.ANTIGO,
    "n/a": EstadoItem.NAO_APLICAVEL,
    "na": EstadoItem.NAO_APLICAVEL,
    "nao aplicavel": EstadoItem.NAO_APLICAVEL,
    "não aplicável": EstadoItem.NAO_APLICAVEL,
    "nao uso": EstadoItem.NAO_APLICAVEL,
    "não uso": EstadoItem.NAO_APLICAVEL,
}


async def _garantir_acesso(db: AsyncIOMotorDatabase, sector_id: str | None, catequista: CatequistaOut) -> None:
    """Admin gere tudo. Sem admin, só o responsável do sector indicado pode
    gerir os itens desse sector — o inventário geral (sem sector) é só admin."""
    if catequista.is_admin:
        return
    if not sector_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Só administradores podem gerir o inventário geral da catequese",
        )
    sector = await db.sectores.find_one({"_id": object_id_or_404(sector_id)})
    if sector is None or sector.get("responsavel_catequista_id") != catequista.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Não és responsável por este sector")


async def _to_out(db: AsyncIOMotorDatabase, doc: dict) -> ItemInventarioOut:
    sector_nome = None
    if doc.get("sector_id"):
        sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])})
        sector_nome = sector["nome"] if sector else None

    return ItemInventarioOut(
        id=str(doc["_id"]),
        nome=doc["nome"],
        sector_id=doc.get("sector_id"),
        sector_nome=sector_nome,
        categoria=doc.get("categoria"),
        quantidade=doc["quantidade"],
        descricao=doc.get("descricao"),
        localizacao=doc.get("localizacao"),
        imagem_url=doc.get("imagem_url"),
        estado=doc.get("estado"),
        criado_em=doc["criado_em"],
        criado_por_nome=doc.get("criado_por_nome"),
        atualizado_em=doc.get("atualizado_em"),
        atualizado_por_nome=doc.get("atualizado_por_nome"),
    )


@router.post("", response_model=ItemInventarioOut, status_code=status.HTTP_201_CREATED)
async def criar_item(
    dados: ItemInventarioCreate,
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    if dados.sector_id and await db.sectores.find_one({"_id": object_id_or_404(dados.sector_id)}) is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="O sector indicado não existe")
    await _garantir_acesso(db, dados.sector_id, catequista)

    doc = {
        "nome": dados.nome.strip(),
        "sector_id": dados.sector_id,
        "categoria": (dados.categoria or "").strip() or None,
        "quantidade": dados.quantidade,
        "descricao": (dados.descricao or "").strip() or None,
        "localizacao": (dados.localizacao or "").strip() or None,
        "imagem_url": (dados.imagem_url or "").strip() or None,
        "estado": dados.estado.value if dados.estado else None,
        "criado_em": datetime.now(timezone.utc),
        "criado_por_nome": catequista.nome,
    }
    result = await db.inventario.insert_one(doc)
    doc["_id"] = result.inserted_id

    await registar(
        db, catequista, AcaoAuditoria.CRIAR, "Inventário", str(result.inserted_id),
        f"Criou item '{doc['nome']}' (qtd: {doc['quantidade']})",
    )

    return await _to_out(db, doc)


@router.get("", response_model=list[ItemInventarioOut])
async def listar_itens(
    db: AsyncIOMotorDatabase = Depends(get_database),
    _: CatequistaOut = Depends(get_current_catequista),
):
    cursor = db.inventario.find().sort("nome", 1)
    return [await _to_out(db, doc) async for doc in cursor]


def _limpar(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    if texto in ("", "-", "--", "#VALUE!"):
        return None
    return texto


def _limpar_imagem(valor) -> str | None:
    texto = _limpar(valor)
    if texto and not texto.lower().startswith("http"):
        return None
    return texto


def _parse_quantidade(valor) -> int:
    try:
        return max(0, int(float(valor)))
    except (TypeError, ValueError):
        return 0


def _parse_estado(valor) -> EstadoItem | None:
    texto = _limpar(valor)
    if not texto:
        return None
    return _MAPA_ESTADO.get(texto.lower())


@router.post("/importar", response_model=ImportacaoResultado)
async def importar_itens(
    sector_id: str | None = Form(None),
    arquivo: UploadFile = File(...),
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    """Importa itens de inventário a partir de um .xlsx com colunas (nesta ordem):
    Nome do Item | Imagem | Descrição | Categoria | Quantidade | Localização | Estado | Observação.
    Só o Nome do Item é obrigatório; as restantes podem ficar em branco."""
    if sector_id and await db.sectores.find_one({"_id": object_id_or_404(sector_id)}) is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="O sector indicado não existe")
    await _garantir_acesso(db, sector_id, catequista)

    conteudo = await arquivo.read()
    try:
        wb = load_workbook(BytesIO(conteudo), data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Não foi possível ler o ficheiro. Confirma que é um .xlsx válido.",
        )

    erros: list[ErroImportacao] = []
    criados = 0
    total = 0

    for indice, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if linha is None or all(v is None for v in linha):
            continue  # linha em branco, ignora sem contar como erro

        total += 1

        nome = _limpar(linha[0]) if len(linha) > 0 else None
        if not nome:
            erros.append(ErroImportacao(linha=indice, motivo="Nome do item em falta"))
            continue

        imagem = _limpar_imagem(linha[1]) if len(linha) > 1 else None
        descricao = _limpar(linha[2]) if len(linha) > 2 else None
        categoria = _limpar(linha[3]) if len(linha) > 3 else None
        quantidade = _parse_quantidade(linha[4]) if len(linha) > 4 else 0
        localizacao = _limpar(linha[5]) if len(linha) > 5 else None
        estado = _parse_estado(linha[6]) if len(linha) > 6 else None
        observacao = _limpar(linha[7]) if len(linha) > 7 else None
        if observacao:
            descricao = f"{descricao}. {observacao}" if descricao else observacao

        doc = {
            "nome": nome,
            "sector_id": sector_id,
            "categoria": categoria,
            "quantidade": quantidade,
            "descricao": descricao,
            "localizacao": localizacao,
            "imagem_url": imagem,
            "estado": estado.value if estado else None,
            "criado_em": datetime.now(timezone.utc),
            "criado_por_nome": catequista.nome,
        }
        await db.inventario.insert_one(doc)
        criados += 1

    await registar(
        db, catequista, AcaoAuditoria.CRIAR, "Inventário", None,
        f"Importou {criados} item(ns) de inventário via Excel" + (f" (sector: {sector_id})" if sector_id else ""),
    )

    return ImportacaoResultado(total_linhas=total, criados=criados, erros=erros)


@router.put("/{item_id}", response_model=ItemInventarioOut)
async def atualizar_item(
    item_id: str,
    dados: ItemInventarioUpdate,
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    oid = object_id_or_404(item_id)
    atual = await db.inventario.find_one({"_id": oid})
    if atual is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item não encontrado")

    update_doc = dados.model_dump(exclude_unset=True)

    sector_id_efetivo = update_doc.get("sector_id", atual.get("sector_id"))
    if "sector_id" in update_doc and update_doc["sector_id"]:
        if await db.sectores.find_one({"_id": object_id_or_404(update_doc["sector_id"])}) is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="O sector indicado não existe")
    await _garantir_acesso(db, sector_id_efetivo, catequista)

    if "nome" in update_doc:
        update_doc["nome"] = update_doc["nome"].strip()
    for campo in ("categoria", "descricao", "localizacao", "imagem_url"):
        if campo in update_doc and update_doc[campo] is not None:
            update_doc[campo] = update_doc[campo].strip() or None
    if "estado" in update_doc and update_doc["estado"] is not None:
        update_doc["estado"] = update_doc["estado"].value

    if not update_doc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nada para atualizar")

    update_doc["atualizado_em"] = datetime.now(timezone.utc)
    update_doc["atualizado_por_nome"] = catequista.nome

    doc = await db.inventario.find_one_and_update(
        {"_id": oid}, {"$set": update_doc}, return_document=ReturnDocument.AFTER
    )

    await registar(db, catequista, AcaoAuditoria.ATUALIZAR, "Inventário", item_id, f"Editou item '{doc['nome']}'")

    return await _to_out(db, doc)


@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
async def apagar_item(
    item_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    oid = object_id_or_404(item_id)
    atual = await db.inventario.find_one({"_id": oid})
    if atual is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item não encontrado")

    await _garantir_acesso(db, atual.get("sector_id"), catequista)

    await db.inventario.delete_one({"_id": oid})

    await registar(db, catequista, AcaoAuditoria.APAGAR, "Inventário", item_id, f"Apagou item '{atual['nome']}'")
