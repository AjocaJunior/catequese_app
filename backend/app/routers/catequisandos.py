from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from bson import ObjectId
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Response, UploadFile, status
from motor.motor_asyncio import AsyncIOMotorDatabase
from openpyxl import load_workbook
from pymongo import ReturnDocument
from pymongo.errors import DuplicateKeyError

from app.core.ano_letivo import obter_ano_letivo_atual
from app.core.auditoria import registar
from app.core.database import get_database
from app.core.deps import get_current_admin, get_current_catequista
from app.core.mongo_utils import object_id_or_404
from app.core.permissoes import garantir_acesso_fase
from app.models.auditoria import AcaoAuditoria
from app.models.caixa import CATEGORIAS_MATRICULA
from app.models.catequista import CatequistaOut
from app.models.pessoa import PessoaOut
from app.models.catequisando import (
    CatequisandoCreate,
    CatequisandoOut,
    CatequisandoUpdate,
    ErroImportacao,
    ImportacaoResultado,
    SituacaoCatequisando,
    TransferenciaRequest,
)
from app.services.pdf_ficha_sacramento import gerar_pdf_ficha_sacramento
from app.services.pdf_guia_transferencia import gerar_pdf_guia_transferencia
from app.services.pdf_lista_catequisandos import gerar_pdf_lista_catequisandos
from app.services.pdf_processo_catequisando import gerar_pdf_processo_catequisando

router = APIRouter(prefix="/catequisandos", tags=["catequisandos"])


async def _fase_nome_ou_erro(db: AsyncIOMotorDatabase, fase_id: str) -> str:
    oid = object_id_or_404(fase_id)
    fase = await db.fases.find_one({"_id": oid})
    if fase is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A fase indicada não existe")
    return fase["nome"]


async def _sector_nome_ou_erro(db: AsyncIOMotorDatabase, sector_id: str) -> str:
    oid = object_id_or_404(sector_id)
    sector = await db.sectores.find_one({"_id": oid})
    if sector is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="O sector indicado não existe")
    return sector["nome"]


def _normalizar_nome(nome: str) -> str:
    return " ".join(nome.strip().lower().split())


async def _garantir_nome_unico(db: AsyncIOMotorDatabase, nome: str, ignorar_id: str | None = None) -> None:
    filtro = {"nome_normalizado": _normalizar_nome(nome)}
    if ignorar_id:
        filtro["_id"] = {"$ne": ObjectId(ignorar_id)}
    existente = await db.catequisandos.find_one(filtro)
    if existente is not None:
        fase_existente = await db.fases.find_one({"_id": object_id_or_404(existente["fase_id"])})
        nome_fase = fase_existente["nome"] if fase_existente else "fase desconhecida"
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Já existe um catequisando registado com o nome '{nome.strip()}' (na {nome_fase})",
        )


def _data_ou_none(valor):
    return valor.date() if valor else None


async def _pessoa_out_ou_none(db: AsyncIOMotorDatabase, pessoa_id: str | None) -> PessoaOut | None:
    if not pessoa_id:
        return None
    p = await db.pessoas.find_one({"_id": object_id_or_404(pessoa_id)})
    if p is None:
        return None
    return PessoaOut(
        id=str(p["_id"]), nome=p["nome"], natural_de=p.get("natural_de"), provincia=p.get("provincia"),
        estado_civil=p.get("estado_civil"), profissao=p.get("profissao"), residencia=p.get("residencia"),
        comunidade=p.get("comunidade"), contacto=p.get("contacto"), criado_em=p["criado_em"],
    )


async def _sector_nome_ou_none(db: AsyncIOMotorDatabase, sector_id: str | None) -> str | None:
    if not sector_id:
        return None
    s = await db.sectores.find_one({"_id": object_id_or_404(sector_id)})
    return s["nome"] if s else None


async def _pessoa_ou_erro(db: AsyncIOMotorDatabase, pessoa_id: str) -> None:
    oid = object_id_or_404(pessoa_id)
    if await db.pessoas.find_one({"_id": oid}) is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A pessoa indicada não existe")


async def _to_out(db: AsyncIOMotorDatabase, doc: dict, fase_nome: str, sector_nome: str | None = None) -> CatequisandoOut:
    return CatequisandoOut(
        id=str(doc["_id"]),
        nome=doc["nome"],
        genero=doc.get("genero"),
        data_nascimento=_data_ou_none(doc.get("data_nascimento")),
        natural_de=doc.get("natural_de"),
        estado_civil=doc.get("estado_civil"),
        profissao=doc.get("profissao"),
        residencia=doc.get("residencia"),
        fase_id=doc["fase_id"],
        fase_nome=fase_nome,
        sector_id=doc.get("sector_id"),
        sector_nome=sector_nome,
        situacao=doc.get("situacao") or SituacaoCatequisando.ATIVO.value,
        data_situacao=_data_ou_none(doc.get("data_situacao")),
        encarregado_nome=doc.get("encarregado_nome"),
        encarregado_contacto=doc.get("encarregado_contacto"),
        encarregado_parentesco=doc.get("encarregado_parentesco"),
        observacoes=doc.get("observacoes"),
        ficha_ano=doc.get("ficha_ano"),
        ficha_numero=doc.get("ficha_numero"),
        pai_id=doc.get("pai_id"),
        pai=await _pessoa_out_ou_none(db, doc.get("pai_id")),
        mae_id=doc.get("mae_id"),
        mae=await _pessoa_out_ou_none(db, doc.get("mae_id")),
        avo_paterno_nome=doc.get("avo_paterno_nome"),
        avo_paterna_nome=doc.get("avo_paterna_nome"),
        avo_materno_nome=doc.get("avo_materno_nome"),
        avo_materna_nome=doc.get("avo_materna_nome"),
        padrinho_id=doc.get("padrinho_id"),
        padrinho=await _pessoa_out_ou_none(db, doc.get("padrinho_id")),
        madrinha_id=doc.get("madrinha_id"),
        madrinha=await _pessoa_out_ou_none(db, doc.get("madrinha_id")),
        eleicao_data=_data_ou_none(doc.get("eleicao_data")),
        escrutinio_1_data=_data_ou_none(doc.get("escrutinio_1_data")),
        escrutinio_2_data=_data_ou_none(doc.get("escrutinio_2_data")),
        escrutinio_3_data=_data_ou_none(doc.get("escrutinio_3_data")),
        baptismo_local=doc.get("baptismo_local"),
        baptismo_data=_data_ou_none(doc.get("baptismo_data")),
        baptismo_certidao_url=doc.get("baptismo_certidao_url"),
        nucleo_id=doc.get("nucleo_id"),
        nucleo_nome=await _sector_nome_ou_none(db, doc.get("nucleo_id")),
        nucleo_comunidade=doc.get("nucleo_comunidade"),
        baptismo_arquidiocese=doc.get("baptismo_arquidiocese"),
        baptismo_livro=doc.get("baptismo_livro"),
        baptismo_numero_registo=doc.get("baptismo_numero_registo"),
        baptismo_pagina=doc.get("baptismo_pagina"),
        primeira_comunhao_data=_data_ou_none(doc.get("primeira_comunhao_data")),
        primeira_comunhao_hora=doc.get("primeira_comunhao_hora"),
        primeira_comunhao_livro=doc.get("primeira_comunhao_livro"),
        primeira_comunhao_numero_registo=doc.get("primeira_comunhao_numero_registo"),
        primeira_comunhao_pagina=doc.get("primeira_comunhao_pagina"),
        crisma_data=_data_ou_none(doc.get("crisma_data")),
        crisma_hora=doc.get("crisma_hora"),
        crisma_livro=doc.get("crisma_livro"),
        crisma_numero_registo=doc.get("crisma_numero_registo"),
        crisma_pagina=doc.get("crisma_pagina"),
        transferencia_numero=doc.get("transferencia_numero"),
        transferencia_ano_guia=doc.get("transferencia_ano_guia"),
        transferencia_ano_frequentado=doc.get("transferencia_ano_frequentado"),
        transferencia_destino_comunidade=doc.get("transferencia_destino_comunidade"),
        transferencia_destino_arquidiocese=doc.get("transferencia_destino_arquidiocese"),
        transferencia_observacao=doc.get("transferencia_observacao"),
        transferencia_data=_data_ou_none(doc.get("transferencia_data")),
        criado_em=doc["criado_em"],
    )


@router.post("", response_model=CatequisandoOut, status_code=status.HTTP_201_CREATED)
async def criar_catequisando(
    dados: CatequisandoCreate,
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    fase_nome = await _fase_nome_ou_erro(db, dados.fase_id)
    sector_nome = await _sector_nome_ou_erro(db, dados.sector_id) if dados.sector_id else None
    if dados.nucleo_id:
        await _sector_nome_ou_erro(db, dados.nucleo_id)
    await _garantir_nome_unico(db, dados.nome)
    for pessoa_id in (dados.pai_id, dados.mae_id, dados.padrinho_id, dados.madrinha_id):
        if pessoa_id:
            await _pessoa_ou_erro(db, pessoa_id)

    doc = dados.model_dump()
    doc["nome"] = doc["nome"].strip()
    doc["nome_normalizado"] = _normalizar_nome(doc["nome"])
    doc["genero"] = dados.genero.value if dados.genero else None
    doc["situacao"] = SituacaoCatequisando.ATIVO.value
    doc["data_situacao"] = None
    doc["criado_em"] = datetime.now(timezone.utc)
    for campo_data in ("data_nascimento", "eleicao_data", "escrutinio_1_data", "escrutinio_2_data",
                       "escrutinio_3_data", "baptismo_data", "primeira_comunhao_data", "crisma_data"):
        if doc.get(campo_data) is not None:
            doc[campo_data] = datetime.combine(doc[campo_data], datetime.min.time())

    try:
        result = await db.catequisandos.insert_one(doc)
    except DuplicateKeyError:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Já existe um catequisando registado com o nome '{doc['nome']}'",
        )
    doc["_id"] = result.inserted_id

    await registar(
        db, catequista, AcaoAuditoria.CRIAR, "Catequisando", str(result.inserted_id),
        f"Criou catequisando '{doc['nome']}' na {fase_nome}",
    )

    return await _to_out(db, doc, fase_nome, sector_nome)


@router.get("", response_model=list[CatequisandoOut])
async def listar_catequisandos(
    fase_id: str | None = Query(None, description="Filtrar por ID da fase"),
    sector_id: str | None = Query(None, description="Filtrar por ID do sector"),
    situacao: SituacaoCatequisando | None = Query(None, description="Filtrar por situação (ativo/crismado)"),
    db: AsyncIOMotorDatabase = Depends(get_database),
    _: CatequistaOut = Depends(get_current_catequista),
):
    filtro: dict = {}
    if fase_id:
        object_id_or_404(fase_id)
        filtro["fase_id"] = fase_id
    if sector_id:
        object_id_or_404(sector_id)
        filtro["sector_id"] = sector_id
    if situacao:
        if situacao == SituacaoCatequisando.ATIVO:
            filtro["situacao"] = {"$nin": [SituacaoCatequisando.CRISMADO.value, SituacaoCatequisando.TRANSFERIDO.value]}
        else:
            filtro["situacao"] = situacao.value

    fases = {str(f["_id"]): f["nome"] async for f in db.fases.find()}
    sectores = {str(s["_id"]): s["nome"] async for s in db.sectores.find()}

    resultado = []
    async for doc in db.catequisandos.find(filtro).sort("nome", 1):
        fase_nome = fases.get(doc["fase_id"], "Fase desconhecida")
        sector_nome = sectores.get(doc.get("sector_id")) if doc.get("sector_id") else None
        resultado.append(await _to_out(db, doc, fase_nome, sector_nome))
    return resultado


@router.get("/pdf")
async def gerar_pdf_lista(
    fase_id: str = Query(...),
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    """PDF com a lista de catequisandos de uma fase (Crismados excluídos —
    já concluíram, não fazem parte da turma ativa)."""
    fase = await garantir_acesso_fase(db, fase_id, catequista)

    catequistas_nomes = []
    for cid in fase.get("catequista_ids", []):
        c = await db.catequistas.find_one({"_id": object_id_or_404(cid)})
        if c:
            catequistas_nomes.append(c["nome"])

    catequisandos = []
    cursor = db.catequisandos.find({
        "fase_id": fase_id,
        "situacao": {"$nin": [SituacaoCatequisando.CRISMADO.value, SituacaoCatequisando.TRANSFERIDO.value]},
    }).sort("nome", 1)
    async for doc in cursor:
        data_nasc = doc.get("data_nascimento")
        catequisandos.append({
            "nome": doc["nome"],
            "data_nascimento": data_nasc,
            "encarregado_contacto": doc.get("encarregado_contacto"),
            "encarregado_parentesco": doc.get("encarregado_parentesco"),
        })

    pdf_bytes = gerar_pdf_lista_catequisandos(fase["nome"], catequistas_nomes, catequisandos)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="lista_{fase["nome"].replace(" ", "_")}.pdf"'},
    )


def _parse_data_nascimento(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    return None


@router.post("/importar", response_model=ImportacaoResultado)
async def importar_catequisandos(
    fase_id: str = Form(...),
    arquivo: UploadFile = File(...),
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    fase = await db.fases.find_one({"_id": object_id_or_404(fase_id)})
    if fase is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="A fase indicada não existe")

    conteudo = await arquivo.read()
    try:
        wb = load_workbook(BytesIO(conteudo), data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Não foi possível ler o ficheiro. Confirma que é um .xlsx válido.")

    erros: list[ErroImportacao] = []
    criados = 0
    total = 0

    for indice, linha in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if linha is None or all(v is None for v in linha):
            continue
        total += 1

        nome = str(linha[0]).strip() if len(linha) > 0 and linha[0] not in (None, "") else None
        if not nome:
            erros.append(ErroImportacao(linha=indice, motivo="Nome em falta"))
            continue

        data_nascimento = _parse_data_nascimento(linha[1]) if len(linha) > 1 else None
        telefone = linha[2] if len(linha) > 2 else None
        contacto = str(telefone).strip() if telefone not in (None, "") else None
        parentesco = str(linha[3]).strip() if len(linha) > 3 and linha[3] not in (None, "") else None

        nome_normalizado = _normalizar_nome(nome)
        if await db.catequisandos.find_one({"nome_normalizado": nome_normalizado}):
            erros.append(ErroImportacao(linha=indice, motivo=f"Já existe um catequisando com o nome '{nome}'"))
            continue

        doc = {
            "nome": nome,
            "nome_normalizado": nome_normalizado,
            "fase_id": fase_id,
            "sector_id": None,
            "data_nascimento": datetime.combine(data_nascimento, datetime.min.time()) if data_nascimento else None,
            "encarregado_nome": None,
            "encarregado_contacto": contacto,
            "encarregado_parentesco": parentesco,
            "observacoes": None,
            "situacao": SituacaoCatequisando.ATIVO.value,
            "criado_em": datetime.now(timezone.utc),
        }
        result = await db.catequisandos.insert_one(doc)
        criados += 1

    await registar(
        db, catequista, AcaoAuditoria.CRIAR, "Catequisando", None,
        f"Importou {criados} catequisando(s) via Excel na fase '{fase['nome']}'",
    )

    return ImportacaoResultado(total_linhas=total, criados=criados, erros=erros)


@router.get("/{catequisando_id}/historico")
async def historico_inscricoes(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    _: CatequistaOut = Depends(get_current_catequista),
):
    oid = object_id_or_404(catequisando_id)
    if await db.catequisandos.find_one({"_id": oid}) is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    resultado = []
    cursor = db.caixa.find({
        "catequisando_id": catequisando_id,
        "categoria": {"$in": list(CATEGORIAS_MATRICULA)},
    }).sort("ano_letivo", -1)
    async for doc in cursor:
        fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])}) if doc.get("fase_id") else None
        resultado.append({
            "ano_letivo": doc.get("ano_letivo"),
            "fase_id": doc.get("fase_id"),
            "fase_nome": fase["nome"] if fase else "Fase desconhecida",
            "categoria": doc["categoria"],
            "data": doc["data"].date().isoformat(),
        })
    return resultado


@router.get("/{catequisando_id}/pdf")
async def gerar_pdf(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one({"_id": oid})
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    await garantir_acesso_fase(db, doc["fase_id"], catequista)

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])}) if doc.get("sector_id") else None

    ano_letivo = await obter_ano_letivo_atual(db)
    inicio_ano = datetime(ano_letivo, 1, 1)
    fim_ano = datetime(ano_letivo + 1, 1, 1)
    presencas = [
        p async for p in db.presencas.find({
            "catequisando_id": catequisando_id,
            "fase_id": doc["fase_id"],
            "data": {"$gte": inicio_ano, "$lt": fim_ano},
        }).sort("data", 1)
    ]

    async def _pessoa_dict(pessoa_id: str | None) -> dict | None:
        if not pessoa_id:
            return None
        return await db.pessoas.find_one({"_id": object_id_or_404(pessoa_id)})

    pessoas_resolvidas = {
        "pai": await _pessoa_dict(doc.get("pai_id")),
        "mae": await _pessoa_dict(doc.get("mae_id")),
        "padrinho": await _pessoa_dict(doc.get("padrinho_id")),
        "madrinha": await _pessoa_dict(doc.get("madrinha_id")),
    }

    pdf_bytes = gerar_pdf_processo_catequisando(doc, fase, sector, presencas, ano_letivo, pessoas_resolvidas)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="ficha_{doc["nome"].replace(" ", "_")}.pdf"'},
    )


@router.put("/{catequisando_id}", response_model=CatequisandoOut)
async def atualizar_catequisando(
    catequisando_id: str,
    dados: CatequisandoUpdate,
    db: AsyncIOMotorDatabase = Depends(get_database),
    admin: CatequistaOut = Depends(get_current_admin),
):
    oid = object_id_or_404(catequisando_id)
    update_doc = dados.model_dump(exclude_unset=True)

    if "nome" in update_doc:
        update_doc["nome"] = update_doc["nome"].strip()
        await _garantir_nome_unico(db, update_doc["nome"], ignorar_id=catequisando_id)
        update_doc["nome_normalizado"] = _normalizar_nome(update_doc["nome"])
    if "genero" in update_doc and update_doc["genero"] is not None:
        update_doc["genero"] = dados.genero.value
    if "situacao" in update_doc and update_doc["situacao"] is not None:
        update_doc["situacao"] = dados.situacao.value
        update_doc["data_situacao"] = datetime.now(timezone.utc)
    for campo_data in ("data_nascimento", "eleicao_data", "escrutinio_1_data", "escrutinio_2_data",
                       "escrutinio_3_data", "baptismo_data", "primeira_comunhao_data", "crisma_data"):
        if campo_data in update_doc and update_doc[campo_data] is not None:
            update_doc[campo_data] = datetime.combine(update_doc[campo_data], datetime.min.time())
    if "fase_id" in update_doc:
        await _fase_nome_ou_erro(db, update_doc["fase_id"])
    if "sector_id" in update_doc and update_doc["sector_id"]:
        await _sector_nome_ou_erro(db, update_doc["sector_id"])
    if "nucleo_id" in update_doc and update_doc["nucleo_id"]:
        await _sector_nome_ou_erro(db, update_doc["nucleo_id"])
    for campo_pessoa in ("pai_id", "mae_id", "padrinho_id", "madrinha_id"):
        if campo_pessoa in update_doc and update_doc[campo_pessoa]:
            await _pessoa_ou_erro(db, update_doc[campo_pessoa])

    if not update_doc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nada para atualizar")

    try:
        doc = await db.catequisandos.find_one_and_update(
            {"_id": oid}, {"$set": update_doc}, return_document=ReturnDocument.AFTER
        )
    except DuplicateKeyError:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Já existe um catequisando registado com esse nome")
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    sector_nome = None
    if doc.get("sector_id"):
        sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])})
        sector_nome = sector["nome"] if sector else None

    await registar(db, admin, AcaoAuditoria.ATUALIZAR, "Catequisando", catequisando_id, f"Editou catequisando '{doc['nome']}'")

    return await _to_out(db, doc, fase["nome"] if fase else "Fase desconhecida", sector_nome)


@router.post("/{catequisando_id}/crismar", response_model=CatequisandoOut)
async def crismar_catequisando(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    admin: CatequistaOut = Depends(get_current_admin),
):
    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one_and_update(
        {"_id": oid},
        {"$set": {"situacao": SituacaoCatequisando.CRISMADO.value, "data_situacao": datetime.now(timezone.utc)}},
        return_document=ReturnDocument.AFTER,
    )
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    sector_nome = None
    if doc.get("sector_id"):
        sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])})
        sector_nome = sector["nome"] if sector else None

    await registar(db, admin, AcaoAuditoria.ATUALIZAR, "Catequisando", catequisando_id, f"Marcou '{doc['nome']}' como Crismado(a)")
    return await _to_out(db, doc, fase["nome"] if fase else "Fase desconhecida", sector_nome)


@router.post("/{catequisando_id}/reativar", response_model=CatequisandoOut)
async def reativar_catequisando(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    admin: CatequistaOut = Depends(get_current_admin),
):
    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one_and_update(
        {"_id": oid},
        {"$set": {"situacao": SituacaoCatequisando.ATIVO.value, "data_situacao": datetime.now(timezone.utc)}},
        return_document=ReturnDocument.AFTER,
    )
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    sector_nome = None
    if doc.get("sector_id"):
        sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])})
        sector_nome = sector["nome"] if sector else None

    await registar(db, admin, AcaoAuditoria.ATUALIZAR, "Catequisando", catequisando_id, f"Reativou '{doc['nome']}' (deixou de estar Crismado)")
    return await _to_out(db, doc, fase["nome"] if fase else "Fase desconhecida", sector_nome)


@router.post("/{catequisando_id}/transferir", response_model=CatequisandoOut)
async def transferir_catequisando(
    catequisando_id: str,
    dados: TransferenciaRequest,
    db: AsyncIOMotorDatabase = Depends(get_database),
    admin: CatequistaOut = Depends(get_current_admin),
):
    """Regista a transferência para outra paróquia/comunidade — mantém o
    registo e o histórico, mas deixa de aparecer nas presenças/pautas ativas
    da fase, tal como um Crismado. A guia (nº/ano) fica guardada para poder
    ser reimpressa depois."""
    oid = object_id_or_404(catequisando_id)

    update_doc = {
        "situacao": SituacaoCatequisando.TRANSFERIDO.value,
        "data_situacao": datetime.now(timezone.utc),
        "transferencia_numero": dados.numero.strip(),
        "transferencia_ano_guia": dados.ano_guia,
        "transferencia_ano_frequentado": dados.ano_frequentado,
        "transferencia_destino_comunidade": dados.destino_comunidade.strip(),
        "transferencia_destino_arquidiocese": dados.destino_arquidiocese.strip(),
        "transferencia_observacao": (dados.observacao or "").strip() or None,
        "transferencia_data": datetime.combine(dados.data_impressao or date.today(), datetime.min.time()),
    }

    doc = await db.catequisandos.find_one_and_update(
        {"_id": oid}, {"$set": update_doc}, return_document=ReturnDocument.AFTER
    )
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    sector_nome = None
    if doc.get("sector_id"):
        sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])})
        sector_nome = sector["nome"] if sector else None

    await registar(
        db, admin, AcaoAuditoria.ATUALIZAR, "Catequisando", catequisando_id,
        f"Transferiu '{doc['nome']}' para '{dados.destino_comunidade.strip()}' "
        f"(guia nº {dados.numero.strip()}/{dados.ano_guia})",
    )
    return await _to_out(db, doc, fase["nome"] if fase else "Fase desconhecida", sector_nome)


@router.get("/{catequisando_id}/transferencia/pdf")
async def gerar_pdf_transferencia(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one({"_id": oid})
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")
    if doc.get("situacao") != SituacaoCatequisando.TRANSFERIDO.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Este catequisando ainda não foi transferido — usa primeiro 'Transferir'.",
        )

    await garantir_acesso_fase(db, doc["fase_id"], catequista)

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    pai = await _pessoa_out_ou_none(db, doc.get("pai_id"))
    mae = await _pessoa_out_ou_none(db, doc.get("mae_id"))
    pessoas = {"pai": pai.model_dump() if pai else None, "mae": mae.model_dump() if mae else None}

    pdf_bytes = gerar_pdf_guia_transferencia(doc, fase, pessoas)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="guia_transferencia_{doc["nome"].replace(" ", "_")}.pdf"'},
    )


_TIPOS_FICHA_SACRAMENTO = {"baptismo", "primeira_comunhao", "crisma"}


@router.get("/{catequisando_id}/ficha-sacramento/pdf")
async def gerar_pdf_ficha_sacramento_endpoint(
    catequisando_id: str,
    tipo: str = Query(..., description="baptismo | primeira_comunhao | crisma"),
    db: AsyncIOMotorDatabase = Depends(get_database),
    catequista: CatequistaOut = Depends(get_current_catequista),
):
    if tipo not in _TIPOS_FICHA_SACRAMENTO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="tipo deve ser 'baptismo', 'primeira_comunhao' ou 'crisma'",
        )

    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one({"_id": oid})
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    await garantir_acesso_fase(db, doc["fase_id"], catequista)

    pai = await _pessoa_out_ou_none(db, doc.get("pai_id"))
    mae = await _pessoa_out_ou_none(db, doc.get("mae_id"))
    padrinho = await _pessoa_out_ou_none(db, doc.get("padrinho_id"))
    madrinha = await _pessoa_out_ou_none(db, doc.get("madrinha_id"))
    nucleo_nome = await _sector_nome_ou_none(db, doc.get("nucleo_id"))

    pessoas = {
        "pai": pai.model_dump() if pai else None,
        "mae": mae.model_dump() if mae else None,
        "padrinho": padrinho.model_dump() if padrinho else None,
        "madrinha": madrinha.model_dump() if madrinha else None,
    }

    pdf_bytes = gerar_pdf_ficha_sacramento(tipo, doc, pessoas, nucleo_nome)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="ficha_{tipo}_{doc["nome"].replace(" ", "_")}.pdf"'},
    )


@router.get("/{catequisando_id}", response_model=CatequisandoOut)
async def obter_catequisando(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    _: CatequistaOut = Depends(get_current_catequista),
):
    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one({"_id": oid})
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    fase = await db.fases.find_one({"_id": object_id_or_404(doc["fase_id"])})
    sector_nome = None
    if doc.get("sector_id"):
        sector = await db.sectores.find_one({"_id": object_id_or_404(doc["sector_id"])})
        sector_nome = sector["nome"] if sector else None

    return await _to_out(db, doc, fase["nome"] if fase else "Fase desconhecida", sector_nome)


@router.delete("/{catequisando_id}", status_code=status.HTTP_204_NO_CONTENT)
async def apagar_catequisando(
    catequisando_id: str,
    db: AsyncIOMotorDatabase = Depends(get_database),
    admin: CatequistaOut = Depends(get_current_admin),
):
    oid = object_id_or_404(catequisando_id)
    doc = await db.catequisandos.find_one({"_id": oid})
    if doc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Catequisando não encontrado")

    await db.catequisandos.delete_one({"_id": oid})

    await registar(db, admin, AcaoAuditoria.APAGAR, "Catequisando", catequisando_id, f"Apagou catequisando '{doc['nome']}'")
